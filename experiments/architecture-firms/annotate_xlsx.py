# -*- coding: utf-8 -*-
"""
Annotate the source firm list (Firms lists_hasi.xlsx) with two columns derived
from the latest scrape run: Status and Note.

  python annotate_xlsx.py

Status:  Scraping | Scraping (0 arch) | Skipped | Not run
Note:    "<ats>: <N> architecture roles (<total> total jobs)" for working firms,
         or the skip reason (from firms.json) for deferred firms.

Re-run after any scrape to refresh the spreadsheet. Picks the most recent
output/<date>/run_summary.json automatically.
"""
import json
import glob
import os
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Firms lists_hasi.xlsx")

run_files = sorted(glob.glob(os.path.join(HERE, "output", "*", "run_summary.json")), key=os.path.getmtime)
if not run_files:
    raise SystemExit("No output/<date>/run_summary.json found — run scrape.ts first.")
summary = json.load(open(run_files[-1], encoding="utf-8"))
firms = json.load(open(os.path.join(HERE, "firms.json"), encoding="utf-8"))

note_by_firm = {f["name"]: f.get("note") for f in firms}
ats_by_firm = {f["name"]: f.get("ats") for f in firms}


def norm(s):
    return "".join(ch.lower() for ch in (s or "") if ch.isalnum())


res = {norm(r["firm"]): r for r in summary["results"]}


def status_note(firm_name):
    r = res.get(norm(firm_name))
    ats = ats_by_firm.get(firm_name) or (r and r.get("detectedAts")) or ""
    if r is None:
        return ("Not run", "")
    if r["status"] == "skipped":
        return ("Skipped", note_by_firm.get(firm_name) or "Skipped")
    arch, total = r["archJobsFound"], r["totalJobsFound"]
    if arch > 0:
        return ("Scraping", f"{ats}: {arch} architecture roles ({total} total jobs)")
    return ("Scraping (0 arch)", f"{ats}: {total} jobs found, 0 architecture. Custom site - may need per-site tuning.")


wb = load_workbook(XLSX)
ws = wb.active

hdr_src = ws.cell(row=1, column=1).font
hdr_font = Font(name=hdr_src.name or "Calibri", bold=True, size=hdr_src.size or 11)
body_font = ws.cell(row=2, column=1).font

status_col, note_col = ws.max_column + 1, ws.max_column + 2
ws.cell(row=1, column=status_col, value="Status").font = hdr_font
ws.cell(row=1, column=note_col, value="Note").font = hdr_font

for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if not name:
        continue
    st, nt = status_note(str(name).strip())
    sc = ws.cell(row=row, column=status_col, value=st)
    nc = ws.cell(row=row, column=note_col, value=nt)
    sc.font = copy(body_font)
    nc.font = copy(body_font)
    sc.alignment = Alignment(vertical="top")
    nc.alignment = Alignment(wrap_text=True, vertical="top")

ws.column_dimensions[ws.cell(row=1, column=status_col).column_letter].width = 18
ws.column_dimensions[ws.cell(row=1, column=note_col).column_letter].width = 70
wb.save(XLSX)
print(f"Annotated {XLSX}\nUsed run: {run_files[-1]}")
